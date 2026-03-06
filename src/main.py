import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

INPUT_DIR = "input"
OUTPUT_DIR = "output"
GROUP_ORDER = ["英希サブスク", "由利子サブスク", "光熱費", "その他"]

GROUP_RULES = {
    "英希サブスク": (
        "モバイル保険",
        "Voicy",
        "Netflix",
        "OPENAI",
        "ANTHROPIC",
        "Netflix(同)",
        "Netflix(同) -東京都 港区",
        "DEEPL",
        "まぐまぐ!",
        "Audible",
        "弥生株式会社",
    ),
    "由利子サブスク": (
        "ZOOM.COM",
        "WEBLIO.CO.JP",
    ),
    "光熱費": (
        "東京ガス",
        "神奈川県営水道上下水道料金",
        "アフラック",
        "パルシステム神奈川〔電力〕",
        "NTT東日本ご利用料金",
        "パルシステム神奈川〔宅配利用代金〕",
        "パルシステム神奈川〔増資等〕",
        "パルシステム神奈川〔共済掛金〕",
        "ドコモご利用料金",
    ),
}

COLUMN_MAP = {
    "利用日": "date",
    "日付": "date",
    "ご利用店名及び商品名": "category",
    "利用店名": "category",
    "利用先": "category",
    "内容": "category",
    "摘要": "category",
    "利用金額": "amount",
    "金額": "amount",
    "支払金額": "amount",
    "支払額": "amount",
}

REQUIRED_COLUMNS = ["date", "category", "amount"]


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns=COLUMN_MAP)


def _find_header_row(path: str, encoding: str) -> int | None:
    with open(path, encoding=encoding, errors="ignore") as f:
        for i, line in enumerate(f):
            if ("利用日" in line) or ("日付" in line):
                return i
    return None


def _read_csv_file(path: str) -> pd.DataFrame | None:
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            header_row = _find_header_row(path, enc)
            if header_row is None:
                continue
            df = pd.read_csv(path, encoding=enc, skiprows=header_row, engine="python")
            print(f"CSV読み込み成功: {path} ({enc})")
            return normalize_columns(df)
        except Exception:
            continue

    print(f"CSV読み込み失敗: {path}")
    return None


def _read_excel_file(path: str) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(path)
        print(f"Excel読み込み成功: {path}")
        return normalize_columns(df)
    except Exception as e:
        print(f"Excel読み込み失敗: {path}")
        print(e)
        return None


def _clean_amount(value) -> float:
    if pd.isna(value):
        return 0.0

    text = str(value).replace(",", "").replace("円", "").replace("¥", "").strip()
    if re.match(r"^\(.+\)$", text):
        text = f"-{text[1:-1]}"

    try:
        return float(text)
    except ValueError:
        return 0.0


def _normalize_category_name(value: str) -> str:
    if pd.isna(value):
        return "未分類"

    text = str(value).strip()
    netflix_aliases = {
        "Netflix(同) -東京都 港区": "Netflix",
        "Netflix(同)": "Netflix",
    }
    return netflix_aliases.get(text, text)


def classify_group(category_text: str) -> str:
    if pd.isna(category_text):
        return "その他"

    text = str(category_text).lower()
    for group_name, keywords in GROUP_RULES.items():
        if any(keyword.lower() in text for keyword in keywords):
            return group_name

    return "その他"


def load_files() -> pd.DataFrame | None:
    if not os.path.isdir(INPUT_DIR):
        print("inputフォルダが見つかりません")
        return None

    dfs: list[pd.DataFrame] = []
    for file_name in sorted(os.listdir(INPUT_DIR)):
        file_path = os.path.join(INPUT_DIR, file_name)

        if file_name.startswith("~$"):
            continue
        if file_name.lower().endswith(".csv"):
            df = _read_csv_file(file_path)
        elif file_name.lower().endswith((".xlsx", ".xls")):
            df = _read_excel_file(file_path)
        else:
            df = None

        if df is not None:
            dfs.append(df)

    if not dfs:
        print("読み込めるファイルがありません")
        return None

    merged = pd.concat(dfs, ignore_index=True)
    merged = normalize_columns(merged).copy()

    for col in REQUIRED_COLUMNS:
        if col not in merged.columns:
            merged[col] = pd.NA

    merged["date"] = pd.to_datetime(merged["date"], errors="coerce")
    merged["category"] = merged["category"].fillna("未分類").apply(
        _normalize_category_name
    )
    merged["amount"] = merged["amount"].apply(_clean_amount)

    merged = merged.dropna(subset=["date"])
    merged = merged[merged["amount"] != 0]

    merged["month"] = merged["date"].dt.strftime("%Y-%m")
    merged["group"] = merged["category"].apply(classify_group)

    merged = merged.sort_values("date").reset_index(drop=True)
    return merged


def create_summary(
    merged: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    monthly_summary = (
        merged.groupby("month", as_index=False)["amount"].sum().sort_values("month")
    )

    category_summary = (
        merged.groupby("category", as_index=False)["amount"]
        .sum()
        .sort_values("amount", ascending=False)
    )

    month_cols = sorted(merged["month"].dropna().unique().tolist())
    group_pivot = merged.pivot_table(
        index=["group", "category"],
        columns="month",
        values="amount",
        aggfunc="sum",
        fill_value=0,
    )
    group_pivot = group_pivot.reindex(columns=month_cols, fill_value=0).reset_index()

    # 月列（YYYY-MM）の最小月・最大月を除外する
    if month_cols:
        min_month = min(month_cols)
        max_month = max(month_cols)
        excluded_months = {min_month, max_month}
        analysis_month_cols = [m for m in month_cols if m not in excluded_months]
        group_pivot = group_pivot.drop(
            columns=[m for m in excluded_months if m in group_pivot.columns],
            errors="ignore",
        )
    else:
        analysis_month_cols = []

    group_pivot["毎月平均額"] = (
        group_pivot[analysis_month_cols].mean(axis=1).round().astype(int)
        if analysis_month_cols
        else 0
    )
    group_pivot["合計額"] = (
        group_pivot[analysis_month_cols].sum(axis=1) if analysis_month_cols else 0.0
    )

    group_summary = group_pivot.rename(
        columns={
            "group": "グループ",
            "category": "費目",
        }
    )
    group_rank = {name: i for i, name in enumerate(GROUP_ORDER)}
    group_summary["_group_rank"] = group_summary["グループ"].map(
        lambda g: group_rank.get(g, len(GROUP_ORDER))
    )
    group_summary = (
        group_summary.sort_values(by=["_group_rank", "費目"])
        .drop(columns=["_group_rank"])
        .reset_index(drop=True)
    )
    # 不要な列が存在する場合は除外する。
    group_summary = group_summary.drop(
        columns=["最初の月", "最後の月"], errors="ignore"
    )

    return monthly_summary, category_summary, group_summary


def _add_charts_to_workbook(output_path: str) -> None:
    wb = load_workbook(output_path)
    if "charts" in wb.sheetnames:
        del wb["charts"]
    charts_ws = wb.create_sheet("charts")
    group_ws = wb["group_summary"]

    charts_ws["A1"] = "光熱費グループ 月別支出推移"

    month_columns: list[tuple[int, str]] = []
    for col_idx in range(3, group_ws.max_column + 1):
        header = group_ws.cell(row=1, column=col_idx).value
        if isinstance(header, str) and re.match(r"^\d{4}-\d{2}$", header):
            month_columns.append((col_idx, header))

    utility_rows: list[int] = []
    for row_idx in range(2, group_ws.max_row + 1):
        group_name = group_ws.cell(row=row_idx, column=1).value
        if group_name == "光熱費":
            utility_rows.append(row_idx)

    # chartsシートに作図用の表を書き出す
    table_header_row = 20
    charts_ws.cell(row=table_header_row, column=1).value = "費目"
    for i, (_, month_label) in enumerate(month_columns, start=2):
        charts_ws.cell(row=table_header_row, column=i).value = month_label

    for out_r, src_r in enumerate(utility_rows, start=table_header_row + 1):
        charts_ws.cell(row=out_r, column=1).value = group_ws.cell(row=src_r, column=2).value
        for out_c, (src_c, _) in enumerate(month_columns, start=2):
            charts_ws.cell(row=out_r, column=out_c).value = group_ws.cell(row=src_r, column=src_c).value

    if month_columns and utility_rows:
        line = LineChart()
        line.title = "光熱費の月別推移"
        line.y_axis.title = "支出金額"
        line.x_axis.title = "年月（YYYY-MM）"

        min_data_col = 1
        max_data_col = 1 + len(month_columns)
        min_data_row = table_header_row + 1
        max_data_row = table_header_row + len(utility_rows)

        data_ref = Reference(
            charts_ws,
            min_col=min_data_col,
            max_col=max_data_col,
            min_row=min_data_row,
            max_row=max_data_row,
        )
        cats_ref = Reference(
            charts_ws,
            min_col=2,
            max_col=max_data_col,
            min_row=table_header_row,
            max_row=table_header_row,
        )
        line.add_data(data_ref, from_rows=True, titles_from_data=True)
        line.set_categories(cats_ref)

        line.height = 10
        line.width = 16
        charts_ws.add_chart(line, "A2")

    wb.save(output_path)


def create_excel_report(
    merged: pd.DataFrame,
    monthly_summary: pd.DataFrame,
    category_summary: pd.DataFrame,
    group_summary: pd.DataFrame,
) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"summary_report_{ts}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="merged_data", index=False)
        monthly_summary.to_excel(writer, sheet_name="monthly_summary", index=False)
        category_summary.to_excel(writer, sheet_name="category_summary", index=False)
        group_summary.to_excel(writer, sheet_name="group_summary", index=False)
        pd.DataFrame({"info": ["charts"]}).to_excel(
            writer, sheet_name="charts", index=False
        )

    _add_charts_to_workbook(output_path)
    _apply_number_formats(output_path)
    return output_path


def _apply_number_formats(output_path: str) -> None:
    wb = load_workbook(output_path)

    merged_ws = wb["merged_data"]
    monthly_ws = wb["monthly_summary"]
    category_ws = wb["category_summary"]
    group_ws = wb["group_summary"]

    def format_column(ws, col_idx: int, start_row: int = 2) -> None:
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "#,##0"

    # merged_data: amount列をヘッダー名で特定
    amount_col_idx = None
    for col_idx in range(1, merged_ws.max_column + 1):
        if merged_ws.cell(row=1, column=col_idx).value == "amount":
            amount_col_idx = col_idx
            break
    if amount_col_idx is not None:
        format_column(merged_ws, amount_col_idx)

    # monthly_summary / category_summary: amount列
    format_column(monthly_ws, 2)
    format_column(category_ws, 2)

    # group_summary: 月列 + 毎月平均額 + 合計額（C列以降）
    for col_idx in range(3, group_ws.max_column + 1):
        format_column(group_ws, col_idx)

    wb.save(output_path)


def main() -> None:
    print("=== Credit Card Analytics Start ===")

    merged = load_files()
    if merged is None or merged.empty:
        print("有効なデータがありません")
        return

    print("列名:", merged.columns.tolist())
    print("読み込み行数:", len(merged))

    monthly_summary, category_summary, group_summary = create_summary(merged)
    output_path = create_excel_report(
        merged=merged,
        monthly_summary=monthly_summary,
        category_summary=category_summary,
        group_summary=group_summary,
    )
    print("Excelレポート作成完了:", output_path)


if __name__ == "__main__":
    main()
